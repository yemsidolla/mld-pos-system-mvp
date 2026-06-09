import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from audit.models import AuditLog
from audit.services import create_audit_log
from catalog.models import Brand, Category, Product, ProductTag, Supplier
from inventory.services import receive_stock

from .models import BatchUploadJob, BatchUploadRow


SCHEMAS = {
    BatchUploadJob.Target.CATEGORIES: {
        "fields": ["name", "description", "is_active"],
        "required": ["name"],
        "sample": {"name": "Food", "description": "Pet food", "is_active": "TRUE"},
    },
    BatchUploadJob.Target.BRANDS: {
        "fields": ["name", "description", "is_active"],
        "required": ["name"],
        "sample": {"name": "Melodu", "description": "House brand", "is_active": "TRUE"},
    },
    BatchUploadJob.Target.SUPPLIERS: {
        "fields": ["name", "contact_person", "phone", "telegram", "address", "notes", "is_active"],
        "required": ["name"],
        "sample": {
            "name": "Pet Wholesale",
            "contact_person": "Sophea",
            "phone": "012345678",
            "telegram": "@supplier",
            "address": "Phnom Penh",
            "notes": "",
            "is_active": "TRUE",
        },
    },
    BatchUploadJob.Target.PRODUCTS: {
        "fields": [
            "product_code",
            "original_barcode",
            "name",
            "category",
            "brand",
            "unit",
            "default_cost_price",
            "default_selling_price",
            "min_stock",
            "description",
            "animal_type",
            "life_stage",
            "tags",
            "is_active",
        ],
        "required": ["product_code", "name"],
        "sample": {
            "product_code": "P001",
            "original_barcode": "8851234567890",
            "name": "Cat Food",
            "category": "Food",
            "brand": "Melodu",
            "unit": "Bag",
            "default_cost_price": "1.50",
            "default_selling_price": "2.50",
            "min_stock": "5",
            "description": "",
            "animal_type": "CAT",
            "life_stage": "ADULT",
            "tags": "Grain Free; Indoor",
            "is_active": "TRUE",
        },
    },
    BatchUploadJob.Target.STOCK_IN: {
        "fields": [
            "product_code",
            "supplier",
            "quantity",
            "expiry_date",
            "actual_unit_cost",
            "landed_unit_cost",
            "selling_price",
            "note",
        ],
        "required": ["product_code", "supplier", "quantity", "expiry_date", "actual_unit_cost", "selling_price"],
        "sample": {
            "product_code": "P001",
            "supplier": "Pet Wholesale",
            "quantity": "10",
            "expiry_date": "2027-06-01",
            "actual_unit_cost": "1.50",
            "landed_unit_cost": "1.75",
            "selling_price": "2.50",
            "note": "Initial stock",
        },
    },
}


def get_schema(target):
    try:
        return SCHEMAS[target]
    except KeyError as exc:
        raise ValidationError("Unsupported upload target.") from exc


def normalize_header(value):
    return str(value or "").strip()


def stringify_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_bool(value, default=True):
    if value in (None, ""):
        return default
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "active"}:
        return True
    if normalized in {"0", "false", "no", "n", "inactive"}:
        return False
    raise ValueError("Use TRUE or FALSE.")


def parse_decimal(value):
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Enter a valid decimal number.") from exc
    if decimal_value < 0:
        raise ValueError("Value cannot be negative.")
    return str(decimal_value)


def parse_positive_int(value):
    try:
        integer = int(str(value).strip())
    except ValueError as exc:
        raise ValueError("Enter a valid whole number.") from exc
    if integer <= 0:
        raise ValueError("Value must be greater than zero.")
    return integer


def parse_non_negative_int(value):
    try:
        integer = int(str(value).strip() or "0")
    except ValueError as exc:
        raise ValueError("Enter a valid whole number.") from exc
    if integer < 0:
        raise ValueError("Value cannot be negative.")
    return integer


def parse_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return date.fromisoformat(str(value).strip()).isoformat()
    except ValueError as exc:
        raise ValueError("Use YYYY-MM-DD.") from exc


# Columns that may be omitted from an uploaded file (kept backward compatible).
OPTIONAL_COLUMNS = {
    BatchUploadJob.Target.STOCK_IN: {"landed_unit_cost"},
    BatchUploadJob.Target.PRODUCTS: {"animal_type", "life_stage", "tags"},
}


def split_tag_names(value):
    raw = str(value or "").replace(";", ",")
    return [name.strip() for name in raw.split(",") if name.strip()]


def validate_headers(target, headers):
    schema = get_schema(target)
    expected = set(schema["fields"])
    provided = {normalize_header(header) for header in headers if normalize_header(header)}
    optional_missing = OPTIONAL_COLUMNS.get(target, set())
    missing = sorted(expected - provided - optional_missing)
    if missing:
        raise ValidationError(f"Missing columns: {', '.join(missing)}")


def parse_csv(uploaded_file):
    uploaded_file.seek(0)
    content = uploaded_file.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    headers = [normalize_header(header) for header in (reader.fieldnames or [])]
    rows = []
    for index, row in enumerate(reader, start=2):
        rows.append((index, {normalize_header(key): stringify_cell(value) for key, value in row.items() if key}))
    return headers, rows


def parse_xlsx(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = [normalize_header(value) for value in next(iterator, [])]
    rows = []
    for index, values in enumerate(iterator, start=2):
        row = {}
        for header, value in zip(headers, values):
            if header:
                row[header] = stringify_cell(value)
        if any(value != "" for value in row.values()):
            rows.append((index, row))
    return headers, rows


def parse_upload_file(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension == ".csv":
        return parse_csv(uploaded_file)
    if extension == ".xlsx":
        return parse_xlsx(uploaded_file)
    raise ValidationError("Only CSV and XLSX files are supported.")


def normalize_row(target, row_data):
    schema = get_schema(target)
    normalized = {}
    errors = []
    warnings = []

    for field in schema["fields"]:
        raw_value = row_data.get(field, "")
        value = stringify_cell(raw_value)
        if field in schema["required"] and value == "":
            errors.append(f"{field}: required")
        normalized[field] = value

    try:
        if target in {BatchUploadJob.Target.CATEGORIES, BatchUploadJob.Target.BRANDS}:
            normalized["is_active"] = parse_bool(normalized.get("is_active", ""), default=True)

        elif target == BatchUploadJob.Target.SUPPLIERS:
            normalized["is_active"] = parse_bool(normalized.get("is_active", ""), default=True)

        elif target == BatchUploadJob.Target.PRODUCTS:
            normalized["is_active"] = parse_bool(normalized.get("is_active", ""), default=True)
            normalized["default_cost_price"] = parse_decimal(normalized.get("default_cost_price") or "0")
            normalized["default_selling_price"] = parse_decimal(normalized.get("default_selling_price") or "0")
            normalized["min_stock"] = parse_non_negative_int(normalized.get("min_stock") or "0")
            if normalized.get("category") and not Category.objects.filter(name=normalized["category"]).exists():
                errors.append("category: category name does not exist")
            if normalized.get("brand") and not Brand.objects.filter(name=normalized["brand"]).exists():
                errors.append("brand: brand name does not exist")
            animal_type = (normalized.get("animal_type") or "").strip().upper()
            if animal_type and animal_type not in {choice for choice, _ in Product.AnimalType.choices}:
                errors.append("animal_type: invalid value")
            normalized["animal_type"] = animal_type
            life_stage = (normalized.get("life_stage") or "").strip().upper()
            if life_stage and life_stage not in {choice for choice, _ in Product.LifeStage.choices}:
                errors.append("life_stage: invalid value")
            normalized["life_stage"] = life_stage
            barcode_value = normalized.get("original_barcode")
            product_code = normalized.get("product_code")
            if barcode_value:
                conflict = Product.objects.filter(original_barcode=barcode_value).exclude(product_code=product_code).exists()
                if conflict:
                    errors.append("original_barcode: already used by another product")
            if product_code and Product.objects.filter(product_code=product_code).exists():
                warnings.append("Existing product will be updated.")

        elif target == BatchUploadJob.Target.STOCK_IN:
            normalized["quantity"] = parse_positive_int(normalized.get("quantity"))
            normalized["expiry_date"] = parse_date(normalized.get("expiry_date"))
            normalized["actual_unit_cost"] = parse_decimal(normalized.get("actual_unit_cost"))
            normalized["landed_unit_cost"] = (
                parse_decimal(normalized.get("landed_unit_cost"))
                if normalized.get("landed_unit_cost")
                else ""
            )
            normalized["selling_price"] = parse_decimal(normalized.get("selling_price"))
            try:
                product = Product.objects.get(product_code=normalized.get("product_code"))
            except Product.DoesNotExist:
                errors.append("product_code: product does not exist")
            else:
                if not product.is_active:
                    errors.append("product_code: product is inactive")
                if not product.original_barcode:
                    errors.append("product_code: product has no original barcode")
            try:
                supplier = Supplier.objects.get(name=normalized.get("supplier"))
            except Supplier.DoesNotExist:
                errors.append("supplier: supplier name does not exist")
            else:
                if not supplier.is_active:
                    errors.append("supplier: supplier is inactive")
    except ValueError as exc:
        errors.append(str(exc))

    return normalized, errors, warnings


def revalidate_row(row):
    normalized, errors, warnings = normalize_row(row.job.target, row.normalized_data)
    row.normalized_data = normalized
    row.validation_errors = errors
    row.warnings = warnings
    row.save(update_fields=["normalized_data", "validation_errors", "warnings", "updated_at"])
    return row


def refresh_job_validations(job):
    rows = list(job.rows.order_by("row_number", "id"))
    for row in rows:
        if not row.is_deleted:
            revalidate_row(row)

    if job.target == BatchUploadJob.Target.PRODUCTS:
        barcode_rows = {}
        for row in rows:
            if row.is_deleted or not row.is_selected:
                continue
            barcode_value = row.normalized_data.get("original_barcode")
            if barcode_value:
                barcode_rows.setdefault(barcode_value, []).append(row)

        for duplicate_rows in barcode_rows.values():
            product_codes = {row.normalized_data.get("product_code") for row in duplicate_rows}
            if len(duplicate_rows) > 1 and len(product_codes) > 1:
                for row in duplicate_rows:
                    errors = list(row.validation_errors)
                    errors.append("original_barcode: duplicate in upload file")
                    row.validation_errors = errors
                    row.save(update_fields=["validation_errors", "updated_at"])

    return list(job.rows.order_by("row_number", "id"))


@transaction.atomic
def create_upload_job(*, target, uploaded_file, uploaded_by):
    headers, parsed_rows = parse_upload_file(uploaded_file)
    validate_headers(target, headers)
    if not parsed_rows:
        raise ValidationError("Uploaded file has no data rows.")

    job = BatchUploadJob.objects.create(
        target=target,
        original_filename=uploaded_file.name,
        uploaded_by=uploaded_by,
    )
    for row_number, raw_data in parsed_rows:
        normalized, errors, warnings = normalize_row(target, raw_data)
        BatchUploadRow.objects.create(
            job=job,
            row_number=row_number,
            raw_data=raw_data,
            normalized_data=normalized,
            validation_errors=errors,
            warnings=warnings,
        )
    refresh_job_validations(job)
    return job


def get_template_csv(target):
    schema = get_schema(target)
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=schema["fields"])
    writer.writeheader()
    writer.writerow(schema["sample"])
    return buffer.getvalue()


def update_upload_row(row, data):
    if row.job.status != BatchUploadJob.Status.PREVIEW:
        raise ValidationError("Committed upload rows cannot be edited.")
    schema = get_schema(row.job.target)
    normalized = {}
    for field in schema["fields"]:
        normalized[field] = stringify_cell(data.get(field, ""))
    row.normalized_data = normalized
    row.raw_data = normalized
    row.is_selected = True
    row.is_deleted = False
    row.save(update_fields=["normalized_data", "raw_data", "is_selected", "is_deleted", "updated_at"])
    refresh_job_validations(row.job)
    return BatchUploadRow.objects.get(pk=row.pk)


def delete_upload_row(row):
    if row.job.status != BatchUploadJob.Status.PREVIEW:
        raise ValidationError("Committed upload rows cannot be deleted.")
    row.is_deleted = True
    row.is_selected = False
    row.save(update_fields=["is_deleted", "is_selected", "updated_at"])
    refresh_job_validations(row.job)
    return BatchUploadRow.objects.get(pk=row.pk)


def _commit_category(data):
    obj, created = Category.objects.update_or_create(
        name=data["name"],
        defaults={"description": data.get("description", ""), "is_active": data.get("is_active", True)},
    )
    return obj, created


def _commit_brand(data):
    obj, created = Brand.objects.update_or_create(
        name=data["name"],
        defaults={"description": data.get("description", ""), "is_active": data.get("is_active", True)},
    )
    return obj, created


def _commit_supplier(data):
    obj, created = Supplier.objects.update_or_create(
        name=data["name"],
        defaults={
            "contact_person": data.get("contact_person", ""),
            "phone": data.get("phone", ""),
            "telegram": data.get("telegram", ""),
            "address": data.get("address", ""),
            "notes": data.get("notes", ""),
            "is_active": data.get("is_active", True),
        },
    )
    return obj, created


def _commit_product(data):
    category = Category.objects.filter(name=data.get("category")).first() if data.get("category") else None
    brand = Brand.objects.filter(name=data.get("brand")).first() if data.get("brand") else None
    obj, created = Product.objects.update_or_create(
        product_code=data["product_code"],
        defaults={
            "original_barcode": data.get("original_barcode") or None,
            "name": data["name"],
            "category": category,
            "brand": brand,
            "unit": data.get("unit") or "Unit",
            "default_cost_price": Decimal(data.get("default_cost_price") or "0"),
            "default_selling_price": Decimal(data.get("default_selling_price") or "0"),
            "min_stock": data.get("min_stock") or 0,
            "description": data.get("description", ""),
            "animal_type": (data.get("animal_type") or "").strip().upper(),
            "life_stage": (data.get("life_stage") or "").strip().upper(),
            "is_active": data.get("is_active", True),
        },
    )
    tag_names = split_tag_names(data.get("tags"))
    if tag_names:
        tags = [ProductTag.objects.get_or_create(name=name)[0] for name in tag_names]
        obj.tags.set(tags)
    return obj, created


def _commit_stock_in(data, user, request=None):
    product = Product.objects.get(product_code=data["product_code"])
    supplier = Supplier.objects.get(name=data["supplier"])
    stock_batch, _movement = receive_stock(
        product=product,
        supplier=supplier,
        quantity=data["quantity"],
        expiry_date=date.fromisoformat(data["expiry_date"]),
        actual_unit_cost=Decimal(data["actual_unit_cost"]),
        landed_unit_cost=Decimal(data["landed_unit_cost"]) if data.get("landed_unit_cost") else None,
        selling_price=Decimal(data["selling_price"]),
        received_by=user,
        request=request,
        note=data.get("note", ""),
    )
    return stock_batch, True


COMMIT_HANDLERS = {
    BatchUploadJob.Target.CATEGORIES: _commit_category,
    BatchUploadJob.Target.BRANDS: _commit_brand,
    BatchUploadJob.Target.SUPPLIERS: _commit_supplier,
    BatchUploadJob.Target.PRODUCTS: _commit_product,
}


@transaction.atomic
def commit_upload_job(*, job, committed_by, request=None):
    job = BatchUploadJob.objects.select_for_update().get(pk=job.pk)
    if job.status != BatchUploadJob.Status.PREVIEW:
        raise ValidationError("This upload job has already been committed.")

    refresh_job_validations(job)
    rows = list(job.rows.select_for_update().order_by("row_number", "id"))

    valid_rows = [row for row in rows if row.can_commit]
    if not valid_rows:
        raise ValidationError("No valid selected rows to commit.")

    summary = {"created": 0, "updated": 0, "skipped": 0, "failed": 0, "rows": len(rows)}
    for row in rows:
        if row.is_deleted or not row.is_selected:
            row.committed_action = "skipped"
            summary["skipped"] += 1
        elif row.validation_errors:
            row.committed_action = "failed"
            summary["failed"] += 1
        else:
            if job.target == BatchUploadJob.Target.STOCK_IN:
                _obj, created = _commit_stock_in(row.normalized_data, committed_by, request=request)
            else:
                _obj, created = COMMIT_HANDLERS[job.target](row.normalized_data)
            row.committed_action = "created" if created else "updated"
            summary[row.committed_action] += 1
        row.save(update_fields=["committed_action", "updated_at"])

    job.mark_committed(summary)
    create_audit_log(
        action=AuditLog.Action.CREATE,
        module="batch_upload",
        user=committed_by,
        request=request,
        object_type="BatchUploadJob",
        object_id=job.pk,
        object_display=f"{job.get_target_display()} {job.original_filename}",
        new_value={"target": job.target, "summary": summary},
    )
    return job
